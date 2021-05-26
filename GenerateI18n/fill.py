#!/usr/bin/env python
#coding=utf-8

# Name:         fill.py
# Desc:         Generate i18n files
# Author:       yutianqi 00290641
# Created:      2021.05.18:59

import configparser
import json
from openpyxl import load_workbook
import os

languageFilePath = ''
srcFile=''
language=''
defaultStr=''

COLUMN_INDEX = {}
DATA = {}

def loadConfiguaration():
    global languageFilePath
    global srcFiles
    global language
    global defaultStr
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    languageFilePath = config.get("prompt","languageFilePath") 
    srcFiles = config.get("prompt","srcFiles") 
    language = config.get("prompt","language") 
    defaultStr = config.get("prompt","defaultStr") 

    if not os.path.isfile(languageFilePath):
        # print('使用本地文件')
        languageFilePath = 'SmartPVMS 7.0国际化信息.xlsx'

def load(filePath):
    if not os.path.isfile(filePath):
        print("找不到国际化资源文件")
        return False

    wb = load_workbook(filePath)
    ws = wb["国际化信息"]
    for row in ws.iter_rows(min_row=1,max_row=1):
        for index, item in enumerate(row):
            COLUMN_INDEX[index] = item.value

    m = {}
    for row in ws.iter_rows(min_row=4,):
        for index, item in enumerate(row):
            if index == 0:
                DATA[item.value]={}
                m = DATA[item.value]
                continue            
            m[COLUMN_INDEX[index]] = item.value
    return True

def updateFile():
    for srcFile in srcFiles.split(';'):
        print('原资源文件：' + srcFile)
        dstFile = srcFile.replace('_zh_CN', '_' + language.replace('-', '_'))
        print('新资源文件：' + dstFile)
        with open(dstFile, 'w+', encoding='utf-8') as f2:
            with open(srcFile, 'r') as f:
                for item in f.readlines():
                    if not item.strip():
                        continue
                    key = item.strip().split('=')[0]
                    if not DATA.get(key):
                        print('  > cannot find key: ' + key)  
                        f2.writelines(key + "=" + defaultStr + '\n')
                        continue

                    f2.writelines(key + "=" + str(DATA[key][language].encode("unicode_escape"))[2:].replace('\\\\', '\\').replace('\\x', '\\u00')[:-1] + '\n')

def start():
    print('''# Name:	  fill.py\n# Desc:	  Generate i18n files\n# Author: yutianqi 00290641''')

def main():
    start()
    loadConfiguaration()
    print('加载国际化文件：' + languageFilePath)
    if not load(languageFilePath):
        return
    updateFile()
    

if __name__ == '__main__': 
    main()
    os.system('pause')

